"""
reader_excel.py - Excel 格式读入（.xlsx/.xls/.xlsm）
自动从 stat_reader.py 拆分生成
"""

from __future__ import annotations
import os
from typing import Any, Optional

import pandas as pd

from .reader_core import (ColumnInfo, ExcelMeta, StatFileResult, _calc_missing_pct, _get_source_type, DATE_ORIGINS, _bilingual)

# ============================================================
# Excel 格式读入（.xlsx/.xls/.xlsm）
# ============================================================

def _fill_merge_range(df: "pd.DataFrame", rlo: int, rhi: int, clo: int, chi: int) -> bool:
    """用合并区域左上角值填充其余单元格。

    行号/列号均为 1-based；1 = 表头行。表头合并（rlo<2）跳过，
    由 pandas 默认保留首列，避免污染数据。返回是否发生了填充。
    """
    if rlo < 2:
        return False
    top_di = rlo - 2
    top_ci = clo - 1
    if not (0 <= top_di < len(df) and 0 <= top_ci < df.shape[1]):
        return False
    top_val = df.iat[top_di, top_ci]
    changed = False
    for r in range(rlo, rhi + 1):
        for c in range(clo, chi + 1):
            di = r - 2
            ci = c - 1
            if 0 <= di < len(df) and 0 <= ci < df.shape[1]:
                try:
                    if df.iat[di, ci] != top_val:
                        df.iat[di, ci] = top_val
                        changed = True
                except Exception:
                    pass
    return changed


def _read_excel(filepath, timestamp, *, format_type, encoding, sheet_name=None, fill_merged_cells: bool = True) -> StatFileResult:
    """读入 Excel 格式文件（.xlsx / .xls）"""
    warnings_list = []
    warnings_list.append(
        "Excel: 默认不含内建统计元数据，仅保留原始数据值和日期基准；"
        "若文件由本技能写出，将尝试从 _col_labels/_val_labels 辅助表还原标签 | "
        "Excel: no built-in statistical metadata by default; labels restored from helper sheets if written by this skill"
    )
    if format_type == "excel_xls":
        xl = pd.ExcelFile(filepath, engine="xlrd")
    else:
        xl = pd.ExcelFile(filepath, engine="openpyxl")
    sheet_names = xl.sheet_names
    # 排除写入时生成的辅助工作表（_col_labels / _val_labels）
    data_sheet_names = [n for n in sheet_names if not n.startswith("_")]

    if len(data_sheet_names) == 0:
        raise ValueError("Excel 文件不包含任何数据工作表（仅有辅助表）")

    selected_sheet = None
    if sheet_name is not None:
        if sheet_name not in sheet_names:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {sheet_names}")
        selected_sheet = sheet_name
    elif len(data_sheet_names) == 1:
        selected_sheet = data_sheet_names[0]
    else:
        sheet_sizes = {}
        for name in data_sheet_names:
            try:
                if format_type == "excel_xls":
                    sheet = xl.book.sheet_by_name(name)
                    sheet_sizes[name] = sheet.nrows
                else:
                    wb = xl.book
                    ws = wb[name]
                    sheet_sizes[name] = ws.max_row
            except Exception:
                sheet_sizes[name] = 0
        selected_sheet = max(sheet_sizes, key=sheet_sizes.get)
        warnings_list.append(
            f"Excel 文件包含 {len(data_sheet_names)} 个数据工作表，已选择数据量最大的 '{selected_sheet}'（约 {sheet_sizes[selected_sheet]} 行）。"
            f"使用 read_all_sheets() 读入全部工作表，或通过 sheet_name 指定工作表。"
        )

    df = pd.read_excel(filepath, sheet_name=selected_sheet)

    # 尝试从辅助工作表还原变量标签 / 值标签（由 _write_excel 生成）
    var_labels: dict = {}
    val_labels: dict = {}
    try:
        col_df = pd.read_excel(filepath, sheet_name="_col_labels")
        for _, row in col_df.iterrows():
            c = row.get("Column")
            l = row.get("Label")
            if pd.notna(c) and str(c).strip() != "":
                var_labels[str(c)] = str(l) if pd.notna(l) else ""
    except Exception:
        pass
    try:
        vl_df = pd.read_excel(filepath, sheet_name="_val_labels")
        for _, row in vl_df.iterrows():
            v = row.get("Variable")
            val = row.get("Value")
            l = row.get("Label")
            if pd.notna(v):
                try:
                    vkey = int(float(val)) if pd.notna(val) and float(val).is_integer() else (float(val) if pd.notna(val) else val)
                except (ValueError, TypeError):
                    vkey = val
                val_labels.setdefault(str(v), {})[vkey] = str(l) if pd.notna(l) else ""
    except Exception:
        pass
    if var_labels or val_labels:
        warnings_list.append(
            f"Excel: 已从辅助表还原 {len(var_labels)} 个变量标签与 {len(val_labels)} 个值标签集 | "
            f"Excel: restored {len(var_labels)} variable labels and {len(val_labels)} value-label sets from helper sheets"
        )
    
    # Extract merge cell ranges from the worksheet + 填充（fill_merged_cells=True）
    merge_cells = []
    filled_any = False
    try:
        if format_type == "excel_xls":
            # xlrd: sheet.merged_cells returns list of (row_low, row_high, col_low, col_high) 0-based 半开区间
            import xlrd
            book = xlrd.open_workbook(filepath, on_demand=True)
            sheet = book.sheet_by_name(selected_sheet)
            for crange in sheet.merged_cells:
                rlo, rhi, clo, chi = crange
                # 转换为 1-based 闭区间（修正原 rhi-1 偏移）
                mc = {"min_row": rlo + 1, "max_row": rhi, "min_col": clo + 1, "max_col": chi}
                merge_cells.append(mc)
                if fill_merged_cells:
                    filled_any |= _fill_merge_range(df, mc["min_row"], mc["max_row"], mc["min_col"], mc["max_col"])
            book.release_resources()
        else:
            # openpyxl: 需用 load_workbook 获取可写 Worksheet，
            # pd.ExcelFile 的 xl.book[...] 是 ReadOnlyWorksheet，无 merged_cells 属性
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
            ws = wb[selected_sheet]
            for crange in ws.merged_cells.ranges:
                mc = {
                    "min_row": crange.min_row, "max_row": crange.max_row,
                    "min_col": crange.min_col, "max_col": crange.max_col,
                }
                merge_cells.append(mc)
                if fill_merged_cells:
                    filled_any |= _fill_merge_range(df, mc["min_row"], mc["max_row"], mc["min_col"], mc["max_col"])
            wb.close()
    except Exception:
        pass

    if filled_any:
        warnings_list.append(_bilingual(
            "Filled merged-cell regions with their top-left value (fill_merged_cells=True) to avoid NaN in other cells",
            "已用合并区域左上角值填充其余合并单元格（fill_merged_cells=True），避免其余单元格为 NaN"
        ))
    
    excel_metadata = {
        "sheet_names": sheet_names,
        "selected_sheet": selected_sheet,
        "total_sheets": len(sheet_names),
    }
    if merge_cells:
        excel_metadata["merge_cell_ranges"] = merge_cells
        warnings_list.append(f"发现 {len(merge_cells)} 个合并单元格区域，已记录在 excel_metadata['merge_cell_ranges'] 中")

    column_report: dict[str, ColumnInfo] = {}
    for col in df.columns:
        column_report[col] = ColumnInfo(
            source_type=_get_source_type(df[col]),
            pandas_dtype=str(df[col].dtype),
            original_label=None,
            has_value_labels=False,
            n_missing=int(df[col].isnull().sum()),
            missing_are_special=False,
            precision_warning=False,
            format_string=None,
            display_width=None,
            storage_width=None,
            measure_level=None,
            alignment=None,
            original_type=None,
        )
        if pd.api.types.is_numeric_dtype(df[col]) or pd.api.types.is_float_dtype(df[col]):
            max_val = df[col].abs().max()
            if pd.notna(max_val) and max_val > 1e15:
                column_report[col]["precision_warning"] = True
                warnings_list.append(f"列 '{col}' 包含 >1e15 的数值，float64 精度可能不足")

    return {
        "dataframe": df,
        "metadata": ExcelMeta(
            file_format="excel", collected_at=timestamp,
            row_count=df.shape[0], column_count=df.shape[1],
            total_missing_pct=_calc_missing_pct(df),
            variable_labels=var_labels, value_labels=val_labels, special_missing={},
            date_origin=DATE_ORIGINS["excel"], file_encoding=None, file_label=None,
            creation_time=None, modification_time=None, notes=[],
            original_variable_types={}, readstat_variable_types={},
            variable_value_labels={}, variable_to_label={},
            missing_user_values={}, missing_ranges={},
            variable_display_width={}, variable_storage_width={},
            variable_measure={}, variable_alignment={},
            column_names_to_labels={}, mr_sets={}, table_name=None,
            excel_metadata=excel_metadata,
        ),
        "warnings": warnings_list,
        "column_report": column_report,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Handlers for additional formats (Matlab, HDF5, Parquet, Feather, ORC, FST, JSON, XML, ODS, HTML, ODM)
# ─────────────────────────────────────────────────────────────────────────────


