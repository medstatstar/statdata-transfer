"""
test_limits_fixes.py — 验证 README「格式限制与解决方案」中已修复项：
  A. Excel 合并单元格填充（fill_merged_cells）
  B. MATLAB v7.3 (HDF5) 回退读取
  C. Parquet 分区目录（多 part 文件合并）
  D. HDF5 数据集/组属性还原变量标签
"""
from __future__ import annotations

import os
import tempfile
from typing import Any

import numpy as np
import pandas as pd
import pytest

import h5py
import pyarrow as pa
import pyarrow.parquet as pq

from scripts import reader_core, writer
from scripts.reader_excel import _read_excel
from scripts.reader_science import (
    _read_matlab_v73,
    _read_parquet_partitioned,
    _collect_attr_labels,
    _extract_full_meta,
)


def _tmp() -> str:
    return tempfile.mkdtemp()


# ---------------------------------------------------------------------------
# A. Excel 合并单元格填充
# ---------------------------------------------------------------------------
def test_excel_merge_cell_fill():
    import openpyxl
    tmp = _tmp()
    xlsx = os.path.join(tmp, "merged.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "id"; ws["B1"] = "name"; ws["C1"] = "grp"; ws["D1"] = "score"
    ws["A2"] = 1; ws["B2"] = "Alice"; ws["C2"] = "G1"; ws["D2"] = 90
    ws["A3"] = 2; ws["B3"] = "Bob";   ws["C3"] = None; ws["D3"] = None
    # 垂直合并：C2:C3 / D2:D3 共享顶部值
    ws.merge_cells("C2:C3")
    ws.merge_cells("D2:D3")
    wb.save(xlsx)

    res = _read_excel(xlsx, "2026-07-12T00:00:00", format_type="excel_xlsx", encoding=None)
    df = res["dataframe"]
    assert df["grp"].tolist() == ["G1", "G1"], "合并单元格 grp 未填充"
    assert df["score"].tolist() == [90, 90], "合并单元格 score 未填充"
    assert any("填充" in w or "Filled" in w for w in res["warnings"]), "缺少填充提示"
    # 合并区域仍记录在元数据
    assert res["metadata"]["excel_metadata"].get("merge_cell_ranges")


def test_excel_merge_cell_fill_disabled():
    import openpyxl
    tmp = _tmp()
    xlsx = os.path.join(tmp, "merged.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    # 两列：id 有 2 行值迫使 DataFrame 为 2 行；grp 合并 A2:A3（仅顶部有值）
    ws["A1"] = "id"; ws["B1"] = "grp"
    ws["A2"] = 1; ws["B2"] = "G1"
    ws["A3"] = 2; ws["B3"] = None
    ws.merge_cells("B2:B3")
    wb.save(xlsx)
    res = _read_excel(xlsx, "2026-07-12T00:00:00", format_type="excel_xlsx",
                      encoding=None, fill_merged_cells=False)
    # 关闭时不填充：非锚点单元格保留 NaN
    assert res["dataframe"]["grp"].isna().any()


# ---------------------------------------------------------------------------
# B. MATLAB v7.3 (HDF5) 回退
# ---------------------------------------------------------------------------
def test_matlab_v73_fallback():
    tmp = _tmp()
    mat = os.path.join(tmp, "v73.mat")
    with h5py.File(mat, "w") as f:
        f.create_dataset("age", data=np.array([10, 20, 30]))
        f.create_dataset("height", data=np.array([150.0, 160.0, 170.0]))
    res = _read_matlab_v73(mat, "2026-07-12T00:00:00", [])
    df = res["dataframe"]
    assert set(df.columns) >= {"age", "height"}
    assert len(df) == 3
    assert res["metadata"]["matlab_metadata"]["mat_file_version"] == "7.3 (HDF5)"


def test_matlab_v73_2d_matrix_columns():
    tmp = _tmp()
    mat = os.path.join(tmp, "v73.mat")
    with h5py.File(mat, "w") as f:
        # 2D 矩阵按列拆分为多列
        f.create_dataset("M", data=np.array([[1, 2, 3], [4, 5, 6]]))  # 2 行 3 列
    res = _read_matlab_v73(mat, "2026-07-12T00:00:00", [])
    df = res["dataframe"]
    assert "M_0" in df.columns and "M_2" in df.columns
    assert len(df) == 2


def test_matlab_v7_still_scipy():
    """回归：普通 v7 .mat 仍走 scipy，不受影响。"""
    import scipy.io
    tmp = _tmp()
    v7 = os.path.join(tmp, "v7.mat")
    scipy.io.savemat(v7, {"x": np.arange(5), "y": np.array([[1, 2], [3, 4]])})
    res = reader_core.read_stat_file(v7)
    assert "x" in res["dataframe"].columns


# ---------------------------------------------------------------------------
# C. Parquet 分区目录
# ---------------------------------------------------------------------------
def test_parquet_partitioned_dir():
    tmp = _tmp()
    pdir = os.path.join(tmp, "pdir")
    os.makedirs(pdir)
    pq.write_table(pa.Table.from_pandas(pd.DataFrame({"a": [1, 2], "b": [3, 4]})),
                   os.path.join(pdir, "part-0.parquet"))
    pq.write_table(pa.Table.from_pandas(pd.DataFrame({"a": [5], "b": [6]})),
                   os.path.join(pdir, "part-1.parquet"))
    res = _read_parquet_partitioned(pdir, "2026-07-12T00:00:00")
    assert len(res["dataframe"]) == 3
    md = res["metadata"]["parquet_metadata"]
    assert md["partitioned"] is True
    assert md["num_files"] == 2


def test_parquet_partitioned_via_dispatch():
    tmp = _tmp()
    pdir = os.path.join(tmp, "pdir")
    os.makedirs(pdir)
    pq.write_table(pa.Table.from_pandas(pd.DataFrame({"a": [1], "b": [2]})),
                   os.path.join(pdir, "part-0.parquet"))
    pq.write_table(pa.Table.from_pandas(pd.DataFrame({"a": [3], "b": [4]})),
                   os.path.join(pdir, "part-1.parquet"))
    res = reader_core.read_stat_file(pdir)
    assert len(res["dataframe"]) == 2


# ---------------------------------------------------------------------------
# D. HDF5 属性标签还原
# ---------------------------------------------------------------------------
def test_hdf5_attr_labels_fallback_path():
    tmp = _tmp()
    h5 = os.path.join(tmp, "labeled.h5")
    with h5py.File(h5, "w") as f:
        f.create_dataset("age", data=np.array([1, 2, 3]))
        f["age"].attrs["label"] = "Age"
        f.create_dataset("weight", data=np.array([4, 5, 6]))
        f["weight"].attrs["description"] = "Body weight"
    # 手工 HDF5 无法被 pd.read_hdf 解析 → 走 h5py 回退路径
    res = reader_core.read_stat_file(h5)
    vl = res["metadata"]["variable_labels"]
    assert vl.get("age") == "Age"
    assert vl.get("weight") == "Body weight"


def test_collect_attr_labels_helper():
    tmp = _tmp()
    h5 = os.path.join(tmp, "x.h5")
    with h5py.File(h5, "w") as f:
        f.create_dataset("col_a", data=np.array([1, 2]))
        f["col_a"].attrs["label"] = "Column A"
        f.create_dataset("col_b", data=np.array([3, 4]))
        f["col_b"].attrs["UNITS"] = "kg"
    with h5py.File(h5, "r") as f:
        found = _collect_attr_labels(f)
    assert found.get("col_a") == "Column A"
    assert found.get("col_b") == "kg"


def test_extract_full_meta_no_crash_on_unrelated_meta():
    """回归：_extract_full_meta 在仅有 pandas 等无关元数据时不应抛错。"""
    out = _extract_full_meta({"pandas": "<schema>"})
    assert out == {}


# ---------------------------------------------------------------------------
# Access .table_name + .twbx 内嵌 Access
# ---------------------------------------------------------------------------
def test_read_stat_file_accepts_table_name():
    """reader.read_stat_file 不拒绝 table_name 关键字参数（避免 regression）。"""
    import inspect
    sig = inspect.signature(reader_core.read_stat_file)
    assert "table_name" in sig.parameters


def test_access_table_name_invalid_raises_clear_error():
    """给定不存在的表名时，收到清晰的 ValueError（不含 TypeError）。"""
    from scripts.reader_legacy import _read_access
    # 用一个目录中的文件代替，仅测错误消息（可换成真实 .mdb 后跳过）
    try:
        _read_access(os.path.join(_tmp(), "no.mdb"), "2026-07-12T00:00:00",
                     table_name="订单")
    except (RuntimeError, ValueError) as e:
        msg = str(e)
        # 不应是 TypeError（参数不接受）
        assert "Argument" not in msg and "unexpected" not in msg and "TypeError" not in msg
